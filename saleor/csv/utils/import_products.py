from django.db import models
from django.db.models.fields import BooleanField, IntegerField, CharField
from django.http import HttpResponse, HttpResponseBadRequest
from django.db.models.fields.related import ManyToManyField, ForeignKey

from openpyxl import load_workbook
from saleor.product.models import Product, ProductVariant, ProductType, Vendor, ProductChannelListing, ProductVariantChannelListing, ProductImage
from django.db import transaction
from saleor.settings import DEFAULT_CURRENCY
from saleor.channel.models import Channel
import datetime
from django.core.exceptions import ObjectDoesNotExist
from decimal import Decimal
from saleor.csv.utils.validators import ColumnValidator

from saleor.product.thumbnails import create_product_thumbnails


# data tiene que ser un diccionario

# variant attributes:
VARIANT_SKU = {"header": "SKU", "attribute": "sku_simple"}
VARIANT_NAME = {"header": "variante", "attribute": "name"}

# products attributes:
PRODUCT_ID = {"header": "id", "attribute": "id_simple"}
PRODUCT_NAME = {"header": "nombre", "attribute": "name"}
PRODUCT_TYPE_NAME = {"header": "categoria", "attribute": "product_type"}
PRODUCT_DESCRIPTION = {"header": "descripcion", "attribute": "description"}

# channel:
CHANNEL_CURRENCY = {"header": "moneda", "attribute": "currency"}

# product channel listing:
CHANNEL_VISIBLE_IN_LISTINGSS = {
    "header": "publicar", "attribute": "visible_in_listings"}
VISIBLE = True
CHANNEL_IS_PUBLISHED = {"header": "publicar", "attribute": "visible_in_listings"}
PUBLISHED = True

# product variant channel listing:
PRICE_AMOUNT = {"header": "precio", "attribute": "price_amount"}
COST_PRICE_AMOUNT = {"header": "costo", "attribute": "cost_price_amount"}


def row_to_object(data, vendor_id):

    id_simple = None
    product_type_name = None
    product_name = None
    sku_simple = None
    variant_name = None
    description = None
    # stocks:
    # peso:
    # imagen:
    # publicar:
    price_amount = None
    cost_price_amount = None
    currency = DEFAULT_CURRENCY

    id_simple = data[PRODUCT_ID['header']]
    product_type_name = data[PRODUCT_TYPE_NAME['header']]
    product_name = data[PRODUCT_NAME['header']]
    sku_simple = data[VARIANT_SKU['header']]
    variant_name = data[VARIANT_NAME['header']]
    description = data[PRODUCT_DESCRIPTION['header']]

    price_amount = Decimal(data[PRICE_AMOUNT['header']]
                           ) if data[PRICE_AMOUNT['header']] else price_amount
    cost_price_amount = Decimal(data[COST_PRICE_AMOUNT['header']]
                                ) if data[COST_PRICE_AMOUNT['header']] else cost_price_amount

    # Retrieve channel:
    channel = Channel.objects.get(currency_code=currency)

    # Retrieve or define variant:
    try:
        product_variant = ProductVariant.objects.get(
            sku_simple=sku_simple, product__vendor_id=vendor_id)
    except ObjectDoesNotExist:
        sku = str(sku_simple) + '00000' + str(vendor_id)
        product_variant = ProductVariant(sku=sku, sku_simple=sku_simple)

    # Retrieve or define product, aslo check if product id is valid:
    try:
        product = product_variant.product
    except ObjectDoesNotExist:
        vendor = Vendor.objects.get(id=vendor_id)
        product = Product(id_simple=id_simple, vendor=vendor)
    product_type = ProductType.objects.get(name=product_type_name)
    product.product_type = product_type
    product.name = product_name

    # product_channel_listing:
    try:
        product_channel = product.channel_listings.get(channel=channel)
    except ObjectDoesNotExist:
        product_channel = ProductChannelListing(channel=channel, product=product)
    product_channel.visible_in_listings = VISIBLE
    product_channel.is_published = PUBLISHED
    product_channel.available_for_purchase = datetime.date.today()
    product_channel.channel = channel

    # product_variant_channel_listing:
    try:
        product_variant_channel = product_variant.channel_listings.get(
            currency=currency)
    except ObjectDoesNotExist:
        product_variant_channel = ProductVariantChannelListing(
            currency=currency, variant=product_variant)
    product_variant_channel.price_amount = price_amount
    product_variant_channel.cost_price_amount = cost_price_amount
    product_variant_channel.channel = channel

    product.save()
    product_variant.product = product
    product_variant.save()
    product_channel.save()
    product_variant_channel.save()

    images = ProductImage.objects.filter(product=product)
    for image in images:
        create_product_thumbnails.delay(image.pk)

    return


def verify_products_from_xlsx(filename):

    errors = []
    for (n, data) in get_row(filename):
        validators = [ColumnValidator]
        for validator in validators:
            error = validator(data, n).validate()
            if error:
                errors.append(error)

    if errors:
        return (False, errors)
    return (True, "archivo valido")


def get_row(filename):
    wb = load_workbook(filename=filename)
    ws = wb.active

    headers = []
    for n, row in enumerate(ws.iter_rows()):
        data = {}
        m = 0
        for cell in row:
            if n == 0:
                headers.append(cell.value)
            else:
                data[headers[m]] = cell.value
                m += 1

        if n != 0:
            yield (n, data)


@transaction.atomic
def insert_products_from_xlsx(filename, vendor_id):
    for (n, data) in get_row(filename):
        if len(data) != 0:
            row_to_object(data, vendor_id)
        try:
            pass
        except Exception as e:
            raise e
    return 'File imported successfully'


def import_products_from_xlsx(filename, vendor_id):
    (success, errors) = verify_products_from_xlsx(filename)
    if not success:
        return (success, errors)
    insert_products_from_xlsx(filename, vendor_id)
    return (True, None)
